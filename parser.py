import links

import requests
from bs4 import BeautifulSoup

import re
import time
from time import gmtime, strftime
from datetime import datetime


import openpyxl
from openpyxl.utils.cell import column_index_from_string as cifs
from openpyxl.styles import Alignment


headers= {"User-Agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.82 Safari/537.36"}
# r = requests.get(links.demo, headers=headers)
# максимальная длинна ссылки до ошибки 414 == 7249

# print(r)
# with open('ex1.html', 'a', encoding="utf-8") as f:
# 	f.write(r.text)

def super_int(string):
	# очищает число от пробелов и символов
	string = str(string)
	for char in string:
		if char not in ['1','2','3','4','5','6','7','8','9','0',".",","]:
			string = string.replace(char, "")
	return int(string.replace(",","."))

def parse_page(doc):
	soup = BeautifulSoup(doc, "html.parser")
	# entrys = soup.findAll("div", class_="search-registry-entrys-block") # хотя правильно entries
	entrys = soup.findAll("div", class_="search-registry-entry-block box-shadow-search-input") # хотя правильно entries
	# print(len(*entrys))
	result = []
	base_url = 'https://zakupki.gov.ru/'

	total_enties = soup.find("div", class_='search-results__total').text
	total_enties = super_int(total_enties)
	print("Заявок по ссылке:",total_enties)

	for e in entrys:
		# Поставщик
		buyer = e.find("div",class_='registry-entry__body-href').find('a')
		# buyer = buyer.get("href")
		buyer_link = base_url+buyer.get("href")
		buyer =	re.sub(r'\s+', ' ', buyer.text)
		# наименование закуки
		subj = e.find("div",class_='registry-entry__body-value').text

		# Номер контракта
		gk = e.find("div", class_='registry-entry__header-mid__number')
		# ссылка на закупку
		gk_link = base_url+gk.find('a').get("href")
		# номер контракта
		gk = re.sub(r'\s+', ' ', gk.text)

		# цена
		price = e.find("div", class_="price-block__value").text
		price = re.sub(r'\s+', '', price)

		dates = e.find('div',class_='data-block mt-auto').findAll('div',class_='data-block__value')
		# дата размещения
		pub_date = dates[0].text
		# дата окончания подачи заявок
		end_date = dates[-1].text

		# ссылка на документы
		doc_link = base_url + e.find("div", class_='href-block mt-auto d-none').find('a').get("href")
		# print(doc_link)

		# print(buyer,subj,gk,gk_link,price)
		result.append((buyer,buyer_link,gk,gk_link,subj,price,pub_date,end_date,doc_link))
	return result

def result_table(pos_list):
	wb = openpyxl.Workbook()
	ws = wb.active
	col_width = {'A': 5,'B': 40,'C': 20, 'D': 40,'E': 15,'F': 12, 'G': 12, 'H': 16}
	for c, w in col_width.items():
		ws.column_dimensions[c].width = w

	cur_row = 1
	ws.cell(cur_row, cifs('A'), value= "№" ).style = "Headline 4"
	ws.cell(cur_row, cifs('B'), value= "Заказчик" ).style = "Headline 4"
	ws.cell(cur_row, cifs('C'), value= "№ ГК" ).style = "Headline 4"
	ws.cell(cur_row, cifs('D'), value= "Закупка" ).style = "Headline 4"
	ws.cell(cur_row, cifs('E'), value= "Цена" ).style = "Headline 4"
	ws.cell(cur_row, cifs('F'), value= "Опубликовано" ).style = "Headline 4"
	ws.cell(cur_row, cifs('G'), value= "Дата окончания" ).style = "Headline 4"
	ws.cell(cur_row, cifs('H'), value= "Документы" ).style = "Headline 4"
	cur_row += 1

	for i,p in enumerate(pos_list,start=1):
		try:
			ws.cell(cur_row, cifs('A'), value= i )

			ws.cell(cur_row, cifs('B'), value= p[0] ).style = 'Hyperlink'
			ws.cell(cur_row, cifs('B') ).hyperlink = p[1]
			# ws.cell(cur_row, cifs('B') ).style.alignment.wrap_text=True

			ws.cell(cur_row, cifs('C'), value= p[2] ).style = 'Hyperlink'
			# ws.cell(cur_row, cifs('C') ).style.alignment.wrap_text=True
			ws.cell(cur_row, cifs('C') ).hyperlink = p[3]

			ws.cell(cur_row, cifs('D'), value= p[4] )
			ws.cell(cur_row, cifs('E'), value= p[5] )
			ws.cell(cur_row, cifs('F'), value= p[6] )
			ws.cell(cur_row, cifs('G'), value= p[7] )

			ws.cell(cur_row, cifs('H'), value= "Документация" ).style = 'Hyperlink'
			ws.cell(cur_row, cifs('H')).hyperlink = p[8]
			cur_row += 1
		except Exception as e:
			print(e,"ошибка в",p)
			ws.cell(cur_row, cifs('A'), value= "Ошибка") 
			cur_row += 1

	today = strftime("%d.%m.%Y ", gmtime())
	file_name = f"Тендеры {str(len(pos_list)) } шт. {today}.xlsx"
	try:
		# output_file = r"G:\Мой диск\! Материально технический отдел\Еженедельный реестр оплат\!Отчёт"+ "\\" + file_name
		output_file = './reports'+ "/" + file_name
		wb.save(output_file)
		print(f"Отчет по {len(pos_list)} тендерам сохранен под названием {file_name}")
	except Exception as e:
		print("Ошибка имени сохраняемого файла :",e)
		output_file = './reports'+ "/"+"Тендеры "+ str(time.time_ns()) +".xlsx"
		wb.save(output_file)
		print(f"Отчет по {len(pos_list)} тендерам сохранен под названием {output_file}")
	finally:
		return output_file

def report_vip():
	search_query = links.vip_orgs()
	pos_list = []
	for i,link in enumerate(search_query):
		pos_list.extend(parse_page(requests.get(link, headers=headers).text))
		time.sleep(0.1)
		print("Ссылка",i)
	return pos_list


def report_okpd(): 
	search_query = links.all_okpd
	pos_list = parse_page(requests.get(search_query, headers=headers).text)
	return pos_list

if __name__ == '__main__':
	output_file = "search_result.xlsx"

	# из локального файла
	# with open(r'ex1.html','r', encoding="utf-8") as f:
	# 	doc_raw = f.read()
	# 	pos_list = parse_page(doc_raw)
	# 	result_table(pos_list,output_file)

	# одна ссылка
	# search_query = links.demo
	# doc_raw =  requests.get(search_query, headers=headers).text
	# pos_list = parse_page(doc_raw)
	# result_table(pos_list,output_file)
	# print(search_query)

	# важные заказчики
	result_table(report_vip())

