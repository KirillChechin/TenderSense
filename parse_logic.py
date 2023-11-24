import links

import requests
from bs4 import BeautifulSoup

import re
import time
from time import gmtime, strftime
from datetime import datetime

import sqlite3
import openpyxl
from openpyxl.utils.cell import column_index_from_string as cifs
from openpyxl.styles import Alignment


# headers= {"User-Agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.82 Safari/537.36"}
headers= {"User-Agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.82 Safari/537.36",'Accept':'text/html'}
# r = requests.get(links.demo, headers=headers)
# максимальная длинна ссылки до ошибки 414 == 7249

# print(r)
# with open('ex1.html', 'a', encoding="utf-8") as f:
# 	f.write(r.text)

def super_int(string):
	"""очищает число от пробелов и символов"""
	string = str(string)
	for char in string:
		if char not in ['1','2','3','4','5','6','7','8','9','0',".",","]:
			string = string.replace(char, "")
	return int(float(string.replace(",",".")))

def parse_page(link,log=True):
	doc = requests.get(link, headers=headers, timeout=15).text
	soup = BeautifulSoup(doc, "html.parser")
	# entrys = soup.findAll("div", class_="search-registry-entrys-block") # хотя правильно entries
	entrys = soup.findAll("div", class_="search-registry-entry-block box-shadow-search-input") # хотя правильно entries
	# print(len(*entrys))
	result = []
	base_url = 'https://zakupki.gov.ru/'

	#если все результаты неуместились на 1-ую страницу, запрашиваем последующие страницы и дополням ими entrys
	total_enties = soup.find("div", class_='search-results__total').text
	total_enties = super_int(total_enties)
	total_enties = min(total_enties,10_000) # потолок 10 000 тендеров, на случай если запрос выдаст 39 млн резуальтатов 
	if links.records_per_page < total_enties:
		total_pages = total_enties//links.records_per_page +1
		print(f"{total_enties} запсией, запрос {total_pages} страниц по {links.records_per_page} на странице")
		for p in range(2,total_pages+1): #начианем со 2 перавая страиц
			page = f'pageNumber={p}&'
			this_link = link + page
			doc = requests.get(this_link, headers=headers, timeout=15).text
			soup = BeautifulSoup(doc, "html.parser")
			entrys.extend(soup.findAll("div", class_="search-registry-entry-block box-shadow-search-input"))
			# print(f"{page}","текущий размер выборки",len(entrys))

	for e in entrys:
		# Поставщик
		buyer = e.find("div",class_='registry-entry__body-href').find('a')
		# buyer = buyer.get("href")
		buyer_link = base_url+buyer.get("href")
		buyer =	re.sub(r'\s+', ' ', buyer.text)
		# наименование закуки
		subj = e.find("div",class_='registry-entry__body-value').text

		# Номер контракта
		gk = e.find("div", class_='registry-entry__header-mid__number')
		# ссылка на закупку
		gk_link = base_url+gk.find('a').get("href")
		# номер контракта
		gk = re.sub(r'\D', '', gk.text)

		# цена
		price = e.find("div", class_="price-block__value").text
		# price = re.sub(r'\s+', '', price)
		price = super_int(price)

		dates = e.find('div',class_='data-block mt-auto').findAll('div',class_='data-block__value')
		# дата размещения
		pub_date = dates[0].text
		# дата окончания подачи заявок
		end_date = dates[-1].text

		# ссылка на документы
		doc_link = base_url + e.find("div", class_='href-block mt-auto d-none').find('a').get("href")

		# Тип аукцина
		tender_type = e.find("div", class_="col-9 p-0 registry-entry__header-top__title text-truncate")
		tender_type = re.sub(r'\s+', ' ', tender_type.text)

		pos = {"buyer":buyer,"buyer_link":buyer_link,"gk":gk,"gk_link":gk_link,"subj":subj,"price":price, "pub_date":pub_date ,"end_date": end_date, "doc_link":doc_link,"tender_type":tender_type}
		
		# собираем доп инфу по каждой заявке
		entry = requests.get(gk_link, headers=headers, timeout=15).text
		soup = BeautifulSoup(entry, "html.parser")
		info_blocks = soup.find_all("div",class_='row blockInfo')
		pos['contract_insurance'] = None
		pos['tender_insurance'] = None
		for block in info_blocks:
			title = block.find('h2')
			title = title.text if title else "NONE"
			# print("просмотр блока",title)
			if "Обеспечение исполнения контракта" in title:
				contract_insurance = block.find_all("span")[:4]
				pos['contract_insurance'] = " /".join(re.sub(r'\s+', ' ', x.text) for x in contract_insurance)
			elif "Обеспечение заявки" in title:
				tender_insurance = block.find_all("span")[:4]
				pos["tender_insurance"] = " /".join(re.sub(r'\s+', ' ', x.text)  for x in tender_insurance)

		result.append(pos)
		links.log_tender(pos["gk"],pos["price"]) if log else None

	return result

def result_table(pos_list):
	wb = openpyxl.Workbook()
	ws = wb.active
	col_width = {'A': 5,'B': 40,'C': 25, 'D': 50,'E': 15,'F': 12, 'G': 12, 'H': 30, 'I': 30, 'J': 30}
	for c, w in col_width.items():
		ws.column_dimensions[c].width = w

	cur_row = 1
	ws.cell(cur_row, cifs('A'), value= "№" ).style = "Headline 4"
	ws.cell(cur_row, cifs('B'), value= "Заказчик" ).style = "Headline 4"
	ws.cell(cur_row, cifs('C'), value= "№ ГК" ).style = "Headline 4"
	ws.cell(cur_row, cifs('D'), value= "Закупка" ).style = "Headline 4"
	ws.cell(cur_row, cifs('E'), value= "Цена" ).style = "Headline 4"
	ws.cell(cur_row, cifs('F'), value= "Опубликовано" ).style = "Headline 4"
	ws.cell(cur_row, cifs('G'), value= "Дата Окончания" ).style = "Headline 4"
	ws.cell(cur_row, cifs('H'), value= "Вид Аукциона" ).style = "Headline 4"
	ws.cell(cur_row, cifs('I'), value= "Обеспечение Контракта" ).style = "Headline 4"
	ws.cell(cur_row, cifs('J'), value= "Обеспечение Заявки" ).style = "Headline 4"
	# ws.cell(cur_row, cifs('H'), value= "Документы" ).style = "Headline 4"
	cur_row += 1
	excluded = 0
	for i,p in enumerate(pos_list,start=1):
		if links.tender_is_stale(p["gk"]):
			# print(p["gk"], "is stale")
			excluded +=1
			continue
		# try:
		ws.cell(cur_row, cifs('A'), value= i )

		ws.cell(cur_row, cifs('B'), value= p["buyer"] ) #.style = 'Hyperlink'
		# ws.cell(cur_row, cifs('B') ).hyperlink = p["buyer_link"]
		ws.cell(cur_row, cifs('B') ).alignment = Alignment(wrap_text=True)

		ws.cell(cur_row, cifs('C'), value= p["gk"] ).style = 'Hyperlink'
		# ws.cell(cur_row, cifs('C') ).style.alignment.wrap_text=True
		ws.cell(cur_row, cifs('C') ).hyperlink = p["gk_link"]

		ws.cell(cur_row, cifs('D'), value= p['subj'] ).alignment = Alignment(wrap_text=True)
		ws.cell(cur_row, cifs('E'), value= p['price'] ).number_format = "### ### ##0.00 ₽" # "₽ #,##0.00"
		ws.cell(cur_row, cifs('F'), value= p['pub_date'] )
		ws.cell(cur_row, cifs('G'), value= p['end_date'] )
		ws.cell(cur_row, cifs('H'), value= p['tender_type'] ).alignment = Alignment(wrap_text=True)

		ws.cell(cur_row, cifs('I'), value= p['contract_insurance'] ).alignment = Alignment(wrap_text=True) if p['contract_insurance'] else None
		ws.cell(cur_row, cifs('J'), value= p['tender_insurance'] ).alignment = Alignment(wrap_text=True) if p['tender_insurance'] else None

		# ws.cell(cur_row, cifs('H'), value= "Документация" ).style = 'Hyperlink'
		# ws.cell(cur_row, cifs('H')).hyperlink = p['doc_link']
		cur_row += 1
		# except Exception as e:
		# 	print(e,"ошибка в",p)
		# 	ws.cell(cur_row, cifs('A'), value= "Ошибка")
		# 	cur_row += 1
	ws.cell(cur_row, cifs('A'), value= f"Исключено из выдачи: {excluded} шт." ).style = "Headline 3"
	cur_row += 1

	today = strftime("%d.%m.%Y ", gmtime())
	file_name = f"Тендеры {str(len(pos_list)-excluded) } шт. {today}.xlsx"
	try:
		# output_file = r"G:\Мой диск\! Материально технический отдел\Еженедельный реестр оплат\!Отчёт"+ "\\" + file_name
		output_file = './reports'+ "/" + file_name
		wb.save(output_file)
		print(f"Отчет по {len(pos_list)} тендерам сохранен под названием {file_name}")
	except Exception as e:
		print("Ошибка имени сохраняемого файла :",e)
		output_file = './reports'+ "/"+"Тендеры "+ str(time.time_ns()) +".xlsx"
		wb.save(output_file)
		print(f"Сохранено под названием {output_file}")
	finally:
		return output_file

def report_orgs():
	search_query = links.vip_orgs()
	search_query = [links.base+links.params+x[1] for x in search_query]
	pos_list = []
	print("просмотр ссылок", len(search_query))
	for i,link in enumerate(search_query):
		query_pos_list = parse_page(link)
		pos_list.extend(query_pos_list)
		# time.sleep(0.1)
	return result_table(pos_list)

def report_okpd():
	search_query = links.all_okpd()
	pos_list = parse_page(search_query)
	return result_table(pos_list)

if __name__ == '__main__':
	output_file = "search_result.xlsx"

	# из локального файла
	# with open(r'ex1.html','r', encoding="utf-8") as f:
	# 	doc_raw = f.read()
	# 	pos_list = parse_page(doc_raw)
	# 	result_table(pos_list,output_file)

	# одна ссылка
	# search_query = links.demo
	# doc_raw =  requests.get(search_query, headers=headers).text
	# pos_list = parse_page(doc_raw)
	# result_table(pos_list,output_file)
	# print(search_query)

	# важные заказчики
	# print(report_orgs())
	print(report_okpd())